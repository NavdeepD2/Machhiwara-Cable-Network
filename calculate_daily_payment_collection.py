import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd
import openpyxl
from datetime import datetime

xlsx_file_name = ""


def browse_file():
    global xlsx_file_name
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls")])
    if file_path:
        try:
            # Read the XLS file
            xls_df = pd.read_excel(file_path)

            # Get the file name without extension
            file_name = file_path.split("/")[-1].split(".")[0]

            # Create a new XLSX file name with "_conv" appended
            xlsx_file_name = file_name + "_conv.xlsx"

            # Save the DataFrame as XLSX
            xls_df.to_excel(xlsx_file_name, index=False)

            # Show a success message
            messagebox.showinfo("Success", f"Conversion complete. File saved as {xlsx_file_name}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")


# Create the main window
root = tk.Tk()
root.title("XLS to XLSX Converter")

# Set the window size to 150x150 pixels
root.geometry("150x150")

# Calculate the center of the screen
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x_center = (screen_width - root.winfo_reqwidth()) / 2
y_center = (screen_height - root.winfo_reqheight()) / 2

# Set the window to open in the center of the screen
root.geometry("+%d+%d" % (x_center, y_center))

# Make the window topmost
root.attributes('-topmost', 1)

# Create a larger button
browse_button = tk.Button(root, text="Browse", command=browse_file, height=3, width=15)
browse_button.pack(pady=20)

# Start the GUI event loop
root.mainloop()
'''
# You can access xlsx_file_name outside the function
if xlsx_file_name:
    print("The converted file name is:", xlsx_file_name)
'''

def replace_employee_names(name):
    if name and name.startswith("01147"):
        return "RachhpalSingh"
    elif name and name.startswith("01149"):
        return "Gurpreet"
    elif name and name.startswith("01143"):
        return "Babbu"
    else:
        return name

def extract_date(date_str):
    if isinstance(date_str, str):
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            return date_obj.date()
        except ValueError:
            return date_str
    else:
        return date_str.date()

def generate_summary_table(file_path2):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path2)
    sheet = workbook.active

    # Create a dictionary to store the collected amount for each employee on each day
    summary_table = {}

    # Iterate through rows (ignore the first row - header) and update employee names while building the summary table
    for row in sheet.iter_rows(min_row=2, values_only=True):
        date, employee_name, amount = row[9], row[10], row[14]  # Columns 10, 11, and 15
        if not employee_name:  # Skip if employee name is empty (None)
            continue
        employee_name = replace_employee_names(employee_name)
        date = extract_date(date)
        if date not in summary_table:
            summary_table[date] = {employee_name: amount}
        else:
            if employee_name not in summary_table[date]:
                summary_table[date][employee_name] = amount
            else:
                summary_table[date][employee_name] += amount

    # Create a new worksheet to display the summary table
    summary_sheet = workbook.create_sheet(title="Summary")

    # Write the headers
    summary_sheet.append(["Date", "Employee Name", "Amount"])

    # Write the summary table data
    for date, employees in summary_table.items():
        for employee, amount in employees.items():
            summary_sheet.append([date, employee, amount])

    # Save the updated workbook
    workbook.save(file_path2)



if __name__ == "__main__":
    generate_summary_table(xlsx_file_name)
